"""
Excel report generator for FinanzIAs.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import datetime


# Color constants
C_BG = "FF0D1117"
C_CARD = "FF161B22"
C_HEADER = "FF21262D"
C_BLUE = "FF58A6FF"
C_GREEN = "FF3FB950"
C_RED = "FFF85149"
C_MUTED = "FF8B949E"
C_TEXT = "FFE6EDF3"
C_BORDER = "FF30363D"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_TEXT, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)


def _border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def generate_portfolio_excel(
    output_path: str,
    portfolio_name: str,
    positions: list,
    prices: dict,
    currency: str = "USD",
    include_tx: bool = True,
) -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Portafolio"
    ws.sheet_view.showGridLines = False

    # Background for all used cells
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=12):
        for cell in row:
            cell.fill = _fill(C_BG)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"FinanzIAs — {portfolio_name}"
    ws["A1"].font = Font(bold=True, color=C_BLUE, size=16, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = _font(color=C_MUTED, size=10)
    ws.row_dimensions[2].height = 20

    # ── Metrics block ────────────────────────────────────────────────────────
    total_invested = sum(p.quantity * p.avg_buy_price for p in positions)
    total_value = 0.0
    for p in positions:
        d = prices.get(p.ticker)
        total_value += (p.quantity * d["price"]) if d else (p.quantity * p.avg_buy_price)
    pl = total_value - total_invested
    pl_pct = (pl / total_invested * 100) if total_invested > 0 else 0.0

    metrics = [
        ("Valor Total", f"{currency} {total_value:,.2f}", C_BLUE),
        ("Invertido", f"{currency} {total_invested:,.2f}", C_MUTED),
        ("P&L", f"{'+'if pl>=0 else ''}{currency} {pl:,.2f}", C_GREEN if pl >= 0 else C_RED),
        ("Rendimiento", f"{pl_pct:+.2f}%", C_GREEN if pl_pct >= 0 else C_RED),
    ]

    start_row = 4
    for i, (label, value, color) in enumerate(metrics):
        col = 2 + i * 2
        label_cell = ws.cell(row=start_row, column=col, value=label)
        label_cell.font = _font(color=C_MUTED, size=9)
        label_cell.alignment = Alignment(horizontal="center")
        label_cell.fill = _fill(C_CARD)

        val_cell = ws.cell(row=start_row + 1, column=col, value=value)
        val_cell.font = Font(bold=True, color=color, size=12, name="Calibri")
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = _fill(C_CARD)

        ws.row_dimensions[start_row].height = 20
        ws.row_dimensions[start_row + 1].height = 28

    # ── Positions table ──────────────────────────────────────────────────────
    headers = [
        "Ticker", "Empresa", "Sector", "Cantidad",
        "P. Compra", "P. Actual", "Var. Hoy %",
        "Invertido", "Valor Actual", "P&L", "P&L %"
    ]
    header_row = 8
    ws.row_dimensions[header_row].height = 22

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=C_MUTED, size=9, name="Calibri")
        cell.fill = _fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    data_row = header_row + 1
    for p in sorted(positions, key=lambda x: x.ticker):
        d = prices.get(p.ticker)
        current = d["price"] if d else None
        change_pct = d.get("change_pct") if d else None
        invested_pos = p.quantity * p.avg_buy_price
        current_val = (p.quantity * current) if current else None
        pos_pl = (current_val - invested_pos) if current_val else None
        pos_pl_pct = ((pos_pl / invested_pos) * 100) if (pos_pl is not None and invested_pos > 0) else None

        row_data = [
            p.ticker,
            p.company_name or p.ticker,
            p.sector or "—",
            p.quantity,
            p.avg_buy_price,
            current,
            change_pct,
            invested_pos,
            current_val,
            pos_pl,
            pos_pl_pct,
        ]

        bg = C_BG if data_row % 2 == 0 else C_CARD
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.alignment = Alignment(
                horizontal="right" if col_idx >= 4 else "left",
                vertical="center"
            )

            if col_idx in (5, 6, 8, 9, 10) and val is not None:
                cell.number_format = f'"{currency}" #,##0.00'
            elif col_idx in (7, 11) and val is not None:
                cell.number_format = '+0.00%;-0.00%'
                cell.value = val / 100 if val is not None else None

            # Color P&L cells
            if col_idx in (10, 11) and val is not None:
                cell.font = Font(
                    bold=True,
                    color=C_GREEN if float(val) >= 0 else C_RED,
                    size=10, name="Calibri"
                )
            else:
                cell.font = _font(size=10)

        ws.row_dimensions[data_row].height = 20
        data_row += 1

    # Column widths
    widths = [10, 28, 18, 12, 14, 14, 12, 14, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Transactions ─────────────────────────────────────────────────
    ws_tx = wb.create_sheet("Transacciones")
    ws_tx.sheet_view.showGridLines = False

    for row in ws_tx.iter_rows(min_row=1, max_row=500, min_col=1, max_col=8):
        for cell in row:
            cell.fill = _fill(C_BG)

    tx_headers = ["Ticker", "Tipo", "Cantidad", "Precio", "Comisiones", "Total", "Fecha", "Notas"]
    for col_idx, h in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color=C_MUTED, size=9, name="Calibri")
        cell.fill = _fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    tx_row = 2
    from database.models import get_session, Transaction, Position as PosModel
    session = get_session()
    try:
        pos_ids = [p.id for p in positions]
        txs = session.query(Transaction).filter(Transaction.position_id.in_(pos_ids)).order_by(Transaction.date.desc()).all()
        pos_map = {p.id: p.ticker for p in positions}
        for tx in txs:
            ticker = pos_map.get(tx.position_id, "?")
            bg = C_BG if tx_row % 2 == 0 else C_CARD
            row_data = [
                ticker, tx.transaction_type, tx.quantity,
                tx.price, tx.fees, tx.total_value,
                tx.date.strftime("%d/%m/%Y") if tx.date else "",
                tx.notes or ""
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_tx.cell(row=tx_row, column=col_idx, value=val)
                cell.fill = _fill(bg)
                cell.font = _font(size=10)
                cell.border = _border()
                cell.alignment = Alignment(
                    horizontal="right" if col_idx in (3, 4, 5, 6) else "left",
                    vertical="center"
                )
            # Color buy/sell
            type_cell = ws_tx.cell(row=tx_row, column=2)
            type_cell.font = Font(
                bold=True,
                color=C_GREEN if tx.transaction_type == "BUY" else C_RED,
                size=10, name="Calibri"
            )
            tx_row += 1
    finally:
        session.close()

    tx_widths = [10, 10, 12, 14, 12, 14, 14, 30]
    for i, w in enumerate(tx_widths, 1):
        ws_tx.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Transaction history (optional) ───────────────────────────────
    if include_tx:
        try:
            from database.models import get_session, Transaction
            session = get_session()
            pos_ids = [p.id for p in positions if hasattr(p, "id")]
            txs = (
                session.query(Transaction)
                .filter(Transaction.position_id.in_(pos_ids))
                .order_by(Transaction.date.desc())
                .limit(500)
                .all()
            )
            pos_map = {p.id: p.ticker for p in positions if hasattr(p, "id")}
            session.expunge_all()
            session.close()

            if txs:
                wt = wb.create_sheet("Transacciones")
                wt.sheet_view.showGridLines = False
                for col in range(1, 8):
                    for row in range(1, len(txs) + 50):
                        wt.cell(row, col).fill = _fill(C_BG)

                tx_headers = ["Fecha", "Ticker", "Tipo", "Cantidad", "Precio", "Comisión", "Total"]
                for ci, h in enumerate(tx_headers, 1):
                    cell = wt.cell(1, ci, h)
                    cell.font = _font(bold=True, color=C_MUTED, size=10)
                    cell.fill = _fill(C_HEADER)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _border()

                for ri, tx in enumerate(txs, 2):
                    ticker = pos_map.get(tx.position_id, "?")
                    total = tx.quantity * tx.price - (tx.fees or 0)
                    row_data = [
                        tx.date.strftime("%d/%m/%Y") if tx.date else "—",
                        ticker,
                        tx.transaction_type,
                        tx.quantity,
                        tx.price,
                        tx.fees or 0.0,
                        total,
                    ]
                    tx_color = C_GREEN if tx.transaction_type == "BUY" else C_RED
                    bg = C_CARD if ri % 2 == 0 else C_BG
                    for ci, val in enumerate(row_data, 1):
                        cell = wt.cell(ri, ci, val)
                        cell.font = _font(color=tx_color if ci == 3 else C_TEXT, size=9)
                        cell.fill = _fill(bg)
                        cell.border = _border()
                        cell.alignment = Alignment(
                            horizontal="right" if ci > 3 else "left",
                            vertical="center"
                        )

                col_ws = [12, 8, 10, 12, 12, 12, 14]
                for ci, w in enumerate(col_ws, 1):
                    wt.column_dimensions[get_column_letter(ci)].width = w
        except Exception as e:
            print(f"[Excel] Transaction sheet error: {e}")

    wb.save(output_path)
    return output_path
